#!/usr/bin/env python3
"""Gera as planilhas Excel formatadas de data/downloads/, uma por seção do
painel (uma para cada setor 2451/2452 onde a seção varia por segmento).
Lê direto dos CSVs de 3_Dados_Tratados_CSV/ — roda depois de build_data.py,
mas não depende do data.json."""

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
SRC_DIR = ROOT / '3_Dados_Tratados_CSV'
OUT_DIR = ROOT / 'data' / 'downloads'

FONT_NAME = 'Arial'
HEADER_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='1F3864')
BODY_FONT = Font(name=FONT_NAME)
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color='595959')
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=13)
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NO_SEP_HINTS = ('ano', 'mes', 'mês', 'cnae', 'codigo', 'código', 'code', 'reporter_code', 'partner_code')


def _read(name, **kwargs):
    defaults = dict(sep=';', decimal=',', encoding='utf-8-sig')
    defaults.update(kwargs)
    return pd.read_csv(SRC_DIR / name, **defaults)


def _sheet_name(name):
    name = re.sub(r'[:\\/?*\[\]]', '-', str(name))
    return name[:31]


def _is_plain_int_col(header):
    h = str(header).lower()
    return any(hint in h for hint in NO_SEP_HINTS)


def _write_df(ws, df, title=None, note=None):
    r = 1
    if title:
        ws.cell(row=r, column=1, value=title).font = TITLE_FONT
        r += 2
    headers = list(df.columns)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=str(h))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    header_row = r
    r += 1
    plain_int_cols = {h for h in headers if _is_plain_int_col(h)}
    for _, row in df.iterrows():
        for c, h in enumerate(headers, start=1):
            v = row[h]
            if pd.isna(v):
                v = None
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                fv = float(v)
                v = int(fv) if fv.is_integer() else fv
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            if isinstance(v, (int, float)):
                cell.number_format = '0' if h in plain_int_cols else '#,##0.##'
        r += 1
    for c, h in enumerate(headers, start=1):
        sample = df[h].head(300)
        max_len = max((len(str(x)) for x in sample if pd.notna(x)), default=10)
        width = max(10, min(42, len(str(h)) + 2, max_len + 2))
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = f'A{header_row + 1}'
    if note:
        nr = r + 1
        ws.cell(row=nr, column=1, value=note).font = NOTE_FONT
        ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=max(1, len(headers)))
    return r


def _new_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _add_sheet(wb, sheet_name, df, title=None, note=None):
    base = _sheet_name(sheet_name)
    name = base
    i = 2
    while name in wb.sheetnames:
        name = _sheet_name(f'{base[:28]}_{i}')
        i += 1
    ws = wb.create_sheet(name)
    _write_df(ws, df, title=title, note=note)
    return ws


def _save(wb, filename):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / filename
    wb.save(path)
    print(f'  data/downloads/{filename}')
    return path


# ---------------------------------------------------------------------------
# 01 — Produção física (nacional, não varia por setor)
# ---------------------------------------------------------------------------
def build_producao_download():
    wb = _new_workbook()
    _add_sheet(wb, 'Aco e gusa bruto', _read('Producao_Fisica_Aco_Gusa_Bruto.csv'),
               title='Produção física — Aço e ferro-gusa (bruto)',
               note='Fonte: Instituto Aço Brasil (IBS).')
    _add_sheet(wb, 'Aco e gusa dessazonalizado', _read('Producao_Fisica_Aco_Gusa_Dessazonalizado.csv'),
               title='Produção física — Aço e ferro-gusa (dessazonalizado)')
    _add_sheet(wb, 'Metalurgia mensal', _read('Producao_Fisica_Mensal_Metalurgia.csv'),
               title='Produção física — Índice mensal da metalurgia',
               note='Fonte: IBGE/PIM-PF.')
    _save(wb, 'Producao_Fisica.xlsx')


# ---------------------------------------------------------------------------
# 02 — Financeiro (PIA, não varia por setor)
# ---------------------------------------------------------------------------
def build_financeiro_download():
    wb = _new_workbook()
    _add_sheet(wb, 'Metalurgia 24', _read('Dados_Financeiros_Metalurgia_24.csv'),
               title='Financeiro — Metalurgia (CNAE 24)',
               note='Fonte: IBGE/PIA-Empresa, 2007-2023.')
    _add_sheet(wb, 'Fundicao 24.5', _read('Dados_Financeiros_Fundicao_24_5.csv'),
               title='Financeiro — Fundição (CNAE 24.5)',
               note='Reúne 2451+2452 sem abertura entre eles. Fonte: IBGE/PIA-Empresa, 2007-2023.')
    _save(wb, 'Financeiro.xlsx')


# ---------------------------------------------------------------------------
# 03 — Emprego formal / RAIS (por setor)
# ---------------------------------------------------------------------------
EMPREGO_FILES = {
    '2451': [
        ('Empregos_RAIS_UF_Ferro_Aco_2451.csv', 'RAIS por UF'),
        ('Empregos_RAIS_Tamanho_Ferro_Aco_2451.csv', 'RAIS por tamanho'),
        ('rais_vinc_fundicao_escolaridade_2451.csv', 'RAIS escolaridade'),
        ('rais_vinc_fundicao_massa_2451.csv', 'RAIS massa salarial'),
        ('rais_vinc_fundicao_ocupacao_2451.csv', 'RAIS ocupacao detalhada'),
        ('rais_vinc_fundicao_ocupacao_agrupada_2451.csv', 'RAIS ocupacao agrupada'),
        ('rais_vinc_fundicao_tempo_emprego_2451.csv', 'RAIS tempo de emprego'),
    ],
    '2452': [
        ('Empregos_RAIS_UF_Nao_Ferrosos_2452.csv', 'RAIS por UF'),
        ('Empregos_RAIS_Tamanho_Nao_Ferrosos_2452.csv', 'RAIS por tamanho'),
        ('rais_vinc_fundicao_escolaridade_2452.csv', 'RAIS escolaridade'),
        ('rais_vinc_fundicao_massa_2452.csv', 'RAIS massa salarial'),
        ('rais_vinc_fundicao_ocupacao_2452.csv', 'RAIS ocupacao detalhada'),
        ('rais_vinc_fundicao_ocupacao_agrupada_2452.csv', 'RAIS ocupacao agrupada'),
        ('rais_vinc_fundicao_tempo_emprego_2452.csv', 'RAIS tempo de emprego'),
    ],
}


def build_emprego_download(cnae, label):
    wb = _new_workbook()
    for fname, sheet in EMPREGO_FILES[cnae]:
        _add_sheet(wb, sheet, _read(fname), title=f'Emprego formal (RAIS) — {label}',
                   note='Fonte: RAIS.')
    _save(wb, f'Emprego_{cnae}.xlsx')


# ---------------------------------------------------------------------------
# 04 — CAGED (arquivos únicos com os dois setores, filtrados por cnae_subclasse)
# ---------------------------------------------------------------------------
def build_caged_download(cnae, label):
    wb = _new_workbook()
    for fname, sheet in [
        ('caged_fundicao_saldo_mensal_uf_cnae.csv', 'CAGED saldo mensal'),
        ('caged_fundicao_salario_mensal_uf_cnae.csv', 'CAGED salario mensal'),
        ('caged_fundicao_tipo_movimentacao_mensal_cnae.csv', 'CAGED tipo movimentacao'),
    ]:
        df = _read(fname)
        df = df[df['cnae_subclasse'].astype(str).str.strip() == cnae].reset_index(drop=True)
        _add_sheet(wb, sheet, df, title=f'CAGED — {label}', note='Fonte: CAGED/Novo CAGED.')
    _save(wb, f'CAGED_{cnae}.xlsx')


# ---------------------------------------------------------------------------
# 05 — Comércio exterior (por setor: NCM de fundição, país/UF, Comtrade)
# ---------------------------------------------------------------------------
COMEX_NCM_FILE = {'2451': 'Comex_2451_NCM7325.csv', '2452': 'Comex_2452_NCM_NaoFerrosos.csv'}
# UF por NCM de fundição, quando já disponível (igual o painel).
COMEX_NCM_UF_FILE = {'2451': 'Comex_2451_NCM7325_UF.csv', '2452': 'Comex_2452_NCM_NaoFerrosos_UF.csv'}


def build_comex_download(cnae, label):
    wb = _new_workbook()
    _add_sheet(wb, 'Balanca (NCM fundicao)', _read(COMEX_NCM_FILE[cnae]),
               title=f'Comércio exterior — Balança comercial (extração NCM) — {label}',
               note='Fonte: Comex Stat/MDIC, extração por NCM específica de fundição (ver painel para a lista de NCMs usada).')
    if cnae in COMEX_NCM_UF_FILE:
        _add_sheet(wb, 'UF (NCM fundicao)', _read(COMEX_NCM_UF_FILE[cnae]),
                   title=f'Comércio exterior — UF de origem, extração NCM — {label}',
                   note='Mesmos NCMs de fundição da aba "Balanca" — usada no ranking por UF do painel.')
    else:
        _add_sheet(wb, 'UF (base ampla)', _read(f'Comex_Exportacao_Importacao_{cnae}.csv'),
                   title=f'Comércio exterior — UF de origem (base ampla) — {label}',
                   note='Base mais ampla que a aba "Balanca" (não restrita ao NCM de fundição) — usada no ranking por UF do painel enquanto não há extração por NCM com essa abertura pra este setor.')
    _add_sheet(wb, 'Pais (base ampla)', _read(f'Comex_Exportacao_Importacao_{cnae}.csv'),
               title=f'Comércio exterior — País parceiro (base ampla) — {label}',
               note='Base mais ampla que a aba "Balanca" (não restrita ao NCM de fundição) — usada no ranking por país do painel.')
    _add_sheet(wb, 'Comtrade mundial', _read(f'Comtrade_Global_Fundicao_{cnae}.csv'),
               title=f'Comércio mundial (Comtrade, proxy por HS) — {label}',
               note='Contexto global — classificação por código HS como proxy do CNAE, não é o número oficial de comércio exterior do Brasil (ver Comex/MDIC pra isso).')
    _save(wb, f'Comercio_Exterior_{cnae}.xlsx')


# ---------------------------------------------------------------------------
# 06 — BNDES (por setor)
# ---------------------------------------------------------------------------
def build_bndes_download(cnae, label):
    wb = _new_workbook()
    _add_sheet(wb, 'Desembolsos', _read(f'BNDES_Desembolsos_{cnae}.csv'),
               title=f'BNDES — Desembolsos — {label}',
               note='Fonte: BNDES, 2002-2026.')
    _save(wb, f'BNDES_{cnae}.xlsx')


# ---------------------------------------------------------------------------
# 07 — DECOM (só ferro e aço, 2451)
# ---------------------------------------------------------------------------
def build_decom_download():
    wb = _new_workbook()
    _add_sheet(wb, 'Processos', _read('decom_fundicao_processos_2451.csv'),
               title='DECOM — Processos de defesa comercial — Ferro e aço (2451)',
               note='Fonte: DECOM/GECEX. Não há processos catalogados para não ferrosos (2452) nesta base.')
    _save(wb, 'DECOM_2451.xlsx')


# ---------------------------------------------------------------------------
# Energia Industrial (view própria, não varia por setor 2451/2452)
# ---------------------------------------------------------------------------
def build_energia_download():
    wb = _new_workbook()
    _add_sheet(wb, 'Estados x divisoes CNAE', _read('energia_industria_transformacao_estados_brasil_2012-2026.csv', decimal='.'),
               title='Energia industrial — Consumo/custo por UF e divisão CNAE (2012-2026)',
               note='Fonte própria (consolidação de dados setoriais de energia).')
    _add_sheet(wb, 'SP detalhado', _read('energia_industria_transformacao_sp_brasil_2012-2026.csv', decimal='.'),
               title='Energia industrial — São Paulo (detalhado)')
    _add_sheet(wb, 'CCEE exato 2451', _read('Consumo_Energia_CCEE_Exato_2451.csv'),
               title='Consumo de energia (CCEE, exato por CNAE) — Ferro e aço (2451)',
               note='Fonte: CCEE, a partir de abr/2024 (mensal).')
    _add_sheet(wb, 'CCEE exato 2452', _read('Consumo_Energia_CCEE_Exato_2452.csv'),
               title='Consumo de energia (CCEE, exato por CNAE) — Não ferrosos (2452)')
    _add_sheet(wb, 'CCEE aproximado metalurgia', _read('Consumo_Energia_CCEE_Metalurgia_e_Produtos_Metal_APROXIMADO.csv'),
               title='Consumo de energia (CCEE, aproximado) — Metalurgia e produtos de metal')
    _save(wb, 'Energia_Industrial.xlsx')


def build_all_downloads():
    print('Gerando planilhas de download (data/downloads/)...')
    build_producao_download()
    build_financeiro_download()
    for cnae, label in [('2451', 'Ferro e aço'), ('2452', 'Não ferrosos')]:
        build_emprego_download(cnae, label)
        build_caged_download(cnae, label)
        build_comex_download(cnae, label)
        build_bndes_download(cnae, label)
    build_decom_download()
    build_energia_download()


if __name__ == '__main__':
    build_all_downloads()
